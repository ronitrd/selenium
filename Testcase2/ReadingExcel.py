import openpyxl

# File-->Workbook-->Sheets-->Rows-->Cells

file="/home/dell/Downloads/file_example_XLSX_10.xlsx"  # (.xls) file doesn't work needs (.xlsx) file
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row  # Count number of rows in an excel sheet
cols=sheet.max_column # Count number of columns in excel sheet

# Reading all the rows and columns from excel sheet
for r in range(1,rows+1):
    for c in range(1,cols+1):
        print(sheet.cell(r,c).value,end="      ")
    print()

