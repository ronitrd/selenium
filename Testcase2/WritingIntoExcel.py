import openpyxl

# # Same Data in every row and column
# file="/home/dell/Downloads/testdata.xlsx"
# workbook=openpyxl.load_workbook(file)
# #sheet=workbook["Sheet1"]
#      #OR
# sheet=workbook.active   #get active sheet from excel
#
# for r in range(1,6):
#     for c in range(1,4):
#         sheet.cell(r,c).value="welcome"
# workbook.save(file)
#========================================================

# Multiple data


file="/home/dell/Downloads/testdata1.xlsx"
workbook=openpyxl.load_workbook(file)

sheet=workbook.active  #OR sheet=workbook["Sheet1"] #get active sheet from excel
sheet.cell(1,1).value=123
sheet.cell(1,2).value="smith"
sheet.cell(1,3).value="engineer"

sheet.cell(2,1).value=567
sheet.cell(2,2).value="john"
sheet.cell(2,3).value="manager"

sheet.cell(3,1).value=789
sheet.cell(3,2).value="david"
sheet.cell(3,3).value="developer"

workbook.save(file)     # Save file after entering the data
